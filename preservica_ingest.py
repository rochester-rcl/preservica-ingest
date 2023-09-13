import os
import os.path
import shutil
import pathlib
import hashlib
import csv
import uuid
import time
import img2pdf
from datetime import datetime
from pyrsistent import thaw
from zipfile import ZipFile
from os.path import basename
from pyPreservica import *
from cleanup_dates import *
from openpyxl import load_workbook
from deepdiff import DeepDiff

#TODO update project path (line 20)
proj_path = '***UPDATE PROJECT PATH***'
proj_log_file = os.path.join(proj_path, 'project_log.txt')

#review bounds of spreadsheet exported from ArchivesSpace NYU Digitization Work Order plugin, espcially maxrow
#TODO update work order spreadsheet file name (line 25) and worksheet name (line 27)
workorder = os.path.join(proj_path, '***UPDATE SPREADSHEET FILE NAME***')
wb = load_workbook(workorder)
ws = wb['***UPDATE WORKSHEET NAME***']
mincol = 2
maxcol = 2
minrow = 2
maxrow = 6
refidcol = 2
aouricol = 3
titlecol = 7
cuidcol = 8
datecol = 12
filecol = 13
prescol = 14

#this function creates the concatenated DPS identifier by merging the organizational ID and the ils/cuid
#TODO update the 'orgid' variable with the relevant value (line 44)
def work_order_cleanup():
    print('----CONCATENATING DPS IDS IN WORK ORDER----')
    orgid = '***UPDATE ORGID***'
    iterrow = 2
    count = 0
    ws.cell(row = 1, column = filecol).value = 'DPS identifier'
    for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maxrow, max_col = maxcol):
        for cell in row:
            filevar = orgid + '_' + ws.cell(row = iterrow, column = cuidcol).value
            ws.cell(row = iterrow, column = filecol).value = filevar
            if ws.cell(row = iterrow, column=datecol).value == None:
                ws.cell(row = iterrow, column=datecol).value = 'undated'
            print('added {}'.format(filevar))
            count += 1
        iterrow += 1
    print('added {} DPS identifiers'.format(count))
    wb.save(workorder)
# work_order_cleanup()

#this function takes the folder containing all the preservation masters and renames to be the "container" folder which will ultimately be used for OPEX incremental ingest
#also creates a "project_log.txt" file to store variables so that an ingest project can be worked on over multiple sessions
def create_container():
    print('----CREATING CONTAINER----')
    project_log_hand = open(proj_log_file, 'a')
    now = datetime.now()
    date_time = now.strftime('%Y-%m-%d_%H-%M-%S')
    project_log_hand.write(date_time + '\n')
    container = 'container_' + date_time
    os.mkdir(os.path.join(proj_path, container))
    project_log_hand.write(container + '\n')
    temp_folder = proj_path
    for file in os.listdir(path = temp_folder):
        if file.endswith('.tif') or file.endswith('.pdf'):
            shutil.move(os.path.join(proj_path, temp_folder, file), os.path.join(proj_path, container, file))
    print('Container directory: {} and moved digital assets into it'.format(container))
    project_log_hand.close()
# create_container()

#TODO use Adobe Bridge and Photoshop or FastStone to convert TIF files into JPEG files (use 90% quality)
#JPEG derivatives will be in 'JPEG' output folder in root of project folder

#This function creates paths to the folders and files and then moves the files to their respective folders.
def folder_ds_files():
    print('----CREATING FOLDER STRUCTURE FOR PRESERVATION MASTERS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_count = 0
    file_count = 0
    path_container = os.path.join(proj_path, container)
    path_JPEG = os.path.join(proj_path, 'JPEG')
    path_list = [path_container, path_JPEG]
    for paths in path_list:
        folder_list = list()
        for file in os.listdir(path = paths):
            file_root = file.split('.')[0].strip()
            if '-' in file_root:
                file_root = file_root.split('-')[0].strip()
            if file_root not in folder_list:
                folder_list.append(file_root)
                print('added {} to folder_list'.format(file_root))
        for file_root in folder_list:
            path_folder = os.path.join(paths, file_root)
            os.mkdir(path_folder)
            print('created {}'.format(path_folder))
            folder_count += 1
        for file in os.listdir(path = paths):
            if '.' not in file:
                continue
            else:
                path_file = os.path.join(paths, file)
                file_prefix = file.split('.')[0].strip()
                if '-' in file_prefix:
                    file_prefix = file_prefix.split('-')[0].strip()
                path_folder = os.path.join(paths, file_prefix, file)
                shutil.move(path_file, path_folder)
                print('moved {} to {}'.format(path_file, path_folder))
                file_count += 1
    print('created {} folders'.format(folder_count))
    print('moved {} files'.format(file_count))   
# folder_ds_files()

#this function takes the constituent JPEG mezzanine files and packages them into one PDF file 
def img_2_pdf():
    print('----CREATING PDFS FROM JPEGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    count = 0
    path_container = os.path.join(proj_path, container)
    path_working = os.path.join(proj_path, 'JPEG')
    for dir in os.listdir(path = path_working):
        subdir_path = os.path.join(path_working, dir)
        imgs = list()
        for file in os.listdir(path = subdir_path):
            file_path = os.path.join(subdir_path, file)
            imgs.append(file_path)
        with open(os.path.join(path_container,  dir, dir + '.pdf'), 'wb') as pdf_convert:
            pdf_convert.write(img2pdf.convert(imgs))
            print('created {}.pdf'.format(dir))
            count += 1
    shutil.rmtree(path_working)
    print('created {} PDF files'.format(count))
# img_2_pdf()

#this function begins to create the PAX structure
#putting PDFs in Representation_Access folders and TIFFs in Representation_Preservation folders
def representation_preservation_access():
    print('----CREATING REPRESENTATION FOLDERS AND MOVING ASSETS INTO THEM----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_count = 0
    file_count = 0
    path_container = os.path.join(proj_path, container)
    rep_pres = 'Representation_Preservation'
    rep_acc = 'Representation_Access'
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        path_pres = os.path.join(proj_path, container, directory, rep_pres)
        path_acc = os.path.join(proj_path, container, directory, rep_acc)
        os.mkdir(path_pres)
        os.mkdir(path_acc)
        folder_count += 2
        for file in os.listdir(path = path_directory):
            path_directoryfile = os.path.join(proj_path, container, directory, file)
            if file == rep_pres or file == rep_acc:
                continue
            elif file.endswith('.pdf'):
                file_name = file.split('.')[0]
                os.mkdir(os.path.join(path_acc, file_name))
                print('created directory: {}'.format(path_acc + '/' + file_name))
                shutil.move(path_directoryfile, os.path.join(path_acc, file_name, file))
                print('moved file: {}'.format(path_acc + '/' + file_name + '/' + file))
                file_count += 1
            else:
                file_name = file.split('.')[0]
                os.mkdir(os.path.join(path_pres, file_name))
                print('created directory: {}'.format(path_pres + '/' + file_name))
                shutil.move(path_directoryfile, os.path.join(path_pres, file_name, file))
                print('moved file: {}'.format(path_pres + '/' + file_name + '/' + file))
                file_count += 1
    print('Created {} Representation directories | Moved {} files into created directories'.format(folder_count, file_count))
# representation_preservation_access()

#TODO run Droid report

#this function stages the "Representation_" folders for each asset inside a new directory,
#then zipes the files into individual PAX objects and deletes the source files
def create_pax():
    print('----CREATING PAX OBJECTS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    project_log_hand.close()
    pax_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        path_paxstage = os.path.join(proj_path, container, directory, 'pax_stage')
        os.mkdir(path_paxstage)
        shutil.move(os.path.join(path_directory, 'Representation_Preservation'), path_paxstage)
        shutil.move(os.path.join(path_directory, 'Representation_Access'), path_paxstage)
        path_directory = os.path.join(proj_path, container, directory)
        zip_dir = pathlib.Path(path_paxstage)
        pax_obj = ZipFile(os.path.join(path_directory, directory + '.zip'), 'w')
        for file_path in zip_dir.rglob("*"):
            pax_obj.write(file_path, arcname = file_path.relative_to(zip_dir))
        pax_obj.close()
        os.rename(os.path.join(path_directory, directory + '.zip'), os.path.join(path_directory, directory + '.pax.zip'))
        pax_count += 1
        shutil.rmtree(path_paxstage)
        print('created {}'.format(str(pax_count) + ': ' + directory + '.pax.zip'))
    print('Created {} PAX objects'.format(pax_count))
# create_pax()

#this function creates the OPEX metadata file that accompanies an individual zipped PAX package
#this function also includes the metadata necessary for ArchivesSpace sync to Preservica
def pax_metadata():
    print('---CREATING METADATA FILES FOR PAX OBJECTS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    dir_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        try:
            pax_hand = open(os.path.join(path_directory, directory + '.pax.zip'), 'rb')
            pax_read = pax_hand.read()
            sha1_checksum = hashlib.sha1(pax_read).hexdigest()
            pax_hand.close()
            iterrow = 2
            for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maxrow, max_col = maxcol):
                for cell in row:
                    cuid = ws.cell(row = iterrow, column = filecol).value
                    if cuid == directory:
                        ref_id = ws.cell(row = iterrow, column = mincol).value
                        title = ws.cell(row = iterrow, column = titlecol).value
                        if '&' in title:
                            title = title.replace('&', 'and')
                        date_full = ws.cell(row = iterrow, column = datecol).value
                        date_formatted = aspace_dates(date_full)
                        display_title = '{title}{date_formatted}'.format(title=title, date_formatted=date_formatted)
                        opex = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
        <opex:Transfer>
            <opex:Fixities>
                <opex:Fixity type="SHA-1" value="{sha1_checksum}"/>
            </opex:Fixities>
        </opex:Transfer>
        <opex:Properties>
            <opex:Title>{title}</opex:Title>
            <opex:Identifiers>
                <opex:Identifier type="code">{ref_id}</opex:Identifier>
            </opex:Identifiers>
        </opex:Properties>
        <opex:DescriptiveMetadata>
            <LegacyXIP xmlns="http://preservica.com/LegacyXIP">
                <AccessionRef>catalogue</AccessionRef>
            </LegacyXIP>
        </opex:DescriptiveMetadata>
    </opex:OPEXMetadata>'''.format(sha1_checksum=sha1_checksum, title=display_title, ref_id=ref_id)
                        filename = directory + '.pax.zip.opex'
                        pax_md_hand = open(os.path.join(path_directory, filename), 'w')
                        pax_md_hand.write(opex)
                        pax_md_hand.close()
                        print('created {}'.format(filename))
                        dir_count += 1
                iterrow += 1
        except:
            print('ERROR: {}'.format(directory))
            project_log_hand = open(proj_log_file, 'a')
            project_log_hand.write(directory + '\n')
            project_log_hand.close()
    print('Created {} OPEX metdata files for individual assets'.format(dir_count))
# pax_metadata()

#this function matches directory names (based on CUID) with archival object numbers from work order spreadsheet
#this metadata is another facet required for ArchivesSpace to Preservica synchronization
def ao_opex_metadata():
    print('----CREATE ARCHIVAL OBJECT OPEX METADATA----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    project_log_hand.close()
    file_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        if directory.startswith('archival_object_'):
            continue
        else:
            iterrow = 2
            for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maxrow, max_col = maxcol):
                for cell in row:
                    cuid = ws.cell(row = iterrow, column = filecol).value
                    if cuid == directory:
                        ao_num_uri = ws.cell(row = iterrow, column = aouricol).value
                        ao_num = 'archival_object_' + ao_num_uri.split('/')[-1].strip()
                        opex = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
        <opex:Properties>
            <opex:Title>{ao_num}</opex:Title>
            <opex:Identifiers>
                <opex:Identifier type="code">{ao_num}</opex:Identifier>
            </opex:Identifiers>
        </opex:Properties>
        <opex:DescriptiveMetadata>
            <LegacyXIP xmlns="http://preservica.com/LegacyXIP">
                <Virtual>false</Virtual>
            </LegacyXIP>
        </opex:DescriptiveMetadata>
    </opex:OPEXMetadata>'''.format(ao_num = ao_num)
                        with open(os.path.join(path_directory, ao_num + '.opex'), 'w') as ao_md:
                            ao_md.write(opex)
                        file_count += 1
                        os.rename(path_directory, os.path.join(path_container, ao_num))
                        print('processed folder metadata in new folder: {}'.format(ao_num))
                iterrow += 1
    print('Created {} archival object metadata files'.format(file_count))
# ao_opex_metadata()

#this function creates the last OPEX metadata file required for the OPEX incremental ingest, for the container folder
#this OPEX file has the folder manifest to ensure that content is ingested properly
def opex_container_metadata():
    print('----CREATE CONTAINER OBJECT OPEX METADATA----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    opex1 = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
    <opex:Transfer>
        <opex:Manifest>
            <opex:Folders>\n'''
    opex2 = ''
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        opex2 += '\t\t\t\t<opex:Folder>' + directory + '</opex:Folder>\n'
    opex3 = '''\t\t\t</opex:Folders>
        </opex:Manifest>
    </opex:Transfer>
</opex:OPEXMetadata>'''
    container_opex_hand = open(os.path.join(proj_path, container, container + '.opex'), 'w')
    container_opex_hand.write(opex1 + opex2 + opex3)
    print('Created OPEX metadata file for {} directory'.format(container))
    container_opex_hand.close()
# opex_container_metadata()

#this function moves the newly ingest assets from the "OPEX Ingest" folder
#to the "pending link" folder to prepare for ArchivesSpace synchronization
def move_opex_aspace():
    print('----MOVING ASSETS TO PENDING LINK----')
    client = EntityAPI()
    opex_folder = client.descendants('db77b64a-64e8-4da2-9645-6f3fe92c3164')
    aspace_folder = client.folder('9370a695-6bd3-441c-8498-982538ee8718')
    count = 0
    for entity in opex_folder:
        client.move(entity, aspace_folder)
        count += 1
        print('moving item {}'.format(str(count)))
        time.sleep(1)
    print('moved {} entities'.format(str(count)))
# move_opex_aspace()

#this function moves the now empty archival_object_###### folders into a newly created container
#folder in "Pending Deletion" to make deletion of the empty folders easier
def move_aspace_trash():
    print('----MOVING EMPTY FOLDERS TO TRASH----')
    client = EntityAPI()
    aspace_folder = client.descendants('9370a695-6bd3-441c-8498-982538ee8718')
    count = 0
    now = datetime.now()
    folder_title = now.strftime('%Y-%m-%d_%H-%M-%S') + '_Deletion'
    new_folder = client.create_folder(folder_title, "container folder to delete AO# folders", 'admin', '6564c3a1-36bf-4b09-ab00-a624ae303f06')
    dest_folder = client.folder(new_folder.reference)
    for entity in aspace_folder:
        test_var = entity.title
        if test_var.startswith('archival_object_'):
            client.move(entity, dest_folder)
            count += 1
            print('moving item {}'.format(str(count)))
            time.sleep(1)
    print('Moved {} folders into the trash'.format(str(count)))
# move_aspace_trash()

#this function uses the "code" identifier to pull the DPS identifier from the work order spreadsheet
#this function also adds the relevant project_id identifier to each asset
#TODO update project_id value (line 410)
def dps_identifier():
    print('----ADDING DPS IDENTIFIERS----')
    client = EntityAPI()
    count = 0
    iterrow = 2
    ws.cell(row = 1, column = prescol).value = 'Preservica UUID'
    for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maxrow, max_col = maxcol):
        for cell in row:
            ref_id = ws.cell(row = iterrow, column = refidcol).value
            cuid = ws.cell(row = iterrow, column = filecol).value
            for ident in filter(only_assets, client.identifier("code", ref_id)):
                asset = client.asset(ident.reference)
                ws.cell(row = iterrow, column = prescol).value = ident.reference
                client.add_identifier(asset, 'dps', cuid)
                client.add_identifier(asset, 'project_id', '***INSERT PROJECT ID***')
            print('adding {} to {}'.format(cuid, ident.reference))
            count += 1
        iterrow += 1
    print('added identifiers to {} digital assets'.format(str(count)))
    wb.save(workorder)
# dps_identifier()

#this function generates PREMIS records for digital assets based on a CSV file
#TODO copy Preservica identifiers over from work order spreadsheet to PREMIS CSV file for first column
#TODO update PREMIS CSV file name in script (line 424)
def premis_generator():
    print('----CREATING PREMIS RECORDS----')
    client = EntityAPI()
    fhand = open('***UPDATE PREMIS CSV FILE NAME***', 'r')
    csv_reader = csv.reader(fhand, delimiter=',')
    count = 0
    for row in csv_reader:
        count += 1
        preservica_uuid = row[0]
        rights_uuid = uuid.uuid4()
        rights_basis = row[1]
        rights_status = row[2]
        rights_jurisdiction = row[3]
        rights_date = row[4]
        rights_note = row[5]
        rights_doc_text = row[6]
        rights_doc_uri = row[7]
        event_1_uuid = uuid.uuid4()
        event_1_type = row[8]
        event_1_datetime = row[9]
        event_1_details = row[10]
        event_1_agent = row[11]
        premis = '''<premis:premis xmlns:premis="http://www.loc.gov/premis/v3" xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://www.loc.gov/premis/v3 https://www.loc.gov/standards/premis/premis.xsd" version="3.0">
        <premis:object xsi:type="premis:intellectualEntity">
            <premis:objectIdentifier>
                <premis:objectIdentifierType>preservica_uuid</premis:objectIdentifierType>
                <premis:objectIdentifierValue>{preservica_uuid}</premis:objectIdentifierValue>
            </premis:objectIdentifier>
        </premis:object>
        <premis:rights>
            <premis:rightsStatement>
                <premis:rightsStatementIdentifier>
                    <premis:rightsStatementIdentifierType>rights_uuid</premis:rightsStatementIdentifierType>
                    <premis:rightsStatementIdentifierValue>{rights_uuid}</premis:rightsStatementIdentifierValue>
                </premis:rightsStatementIdentifier>
                <premis:rightsBasis authority="rightsBasis" authorityURI="http://id.loc.gov/vocabulary/preservation/rightsBasis" valueURI="http://id.loc.gov/vocabulary/preservation/rightsBasis/cop">{rights_basis}</premis:rightsBasis>
                <premis:copyrightInformation>
                    <premis:copyrightStatus>{rights_status}</premis:copyrightStatus>
                    <premis:copyrightJurisdiction>{rights_jurisdiction}</premis:copyrightJurisdiction>
                    <premis:copyrightStatusDeterminationDate>{rights_date}</premis:copyrightStatusDeterminationDate>
                    <premis:copyrightNote>{rights_note}</premis:copyrightNote>
                    <premis:copyrightDocumentationIdentifier>
                        <premis:copyrightDocumentationIdentifierType>{rights_doc_text}</premis:copyrightDocumentationIdentifierType>
                        <premis:copyrightDocumentationIdentifierValue>{rights_doc_uri}</premis:copyrightDocumentationIdentifierValue>
                    </premis:copyrightDocumentationIdentifier>
                </premis:copyrightInformation>
                <premis:linkingObjectIdentifier>
                    <premis:linkingObjectIdentifierType>preservica_uuid</premis:linkingObjectIdentifierType>
                    <premis:linkingObjectIdentifierValue>{preservica_uuid}</premis:linkingObjectIdentifierValue>
                </premis:linkingObjectIdentifier>
            </premis:rightsStatement>
        </premis:rights>
        <premis:event>
            <premis:eventIdentifier>
                <premis:eventIdentifierType>event_uuid</premis:eventIdentifierType>
                <premis:eventIdentifierValue>{event_1_uuid}</premis:eventIdentifierValue>
            </premis:eventIdentifier>
            <premis:eventType>{event_1_type}</premis:eventType>
            <premis:eventDateTime>{event_1_datetime}</premis:eventDateTime>
            <premis:eventDetailInformation>
                <premis:eventDetail>{event_1_details}</premis:eventDetail>
            </premis:eventDetailInformation>
            <premis:linkingAgentIdentifier>
                <premis:linkingAgentIdentifierType>local</premis:linkingAgentIdentifierType>
                <premis:linkingAgentIdentifierValue>{event_1_agent}</premis:linkingAgentIdentifierValue>
                <premis:linkingAgentRole authority="eventRelatedAgentRole" authorityURI="http://id.loc.gov/vocabulary/preservation/eventRelatedAgentRole" valueURI="http://id.loc.gov/vocabulary/preservation/eventRelatedAgentRole/imp">implementer</premis:linkingAgentRole>
            </premis:linkingAgentIdentifier>
            <premis:linkingObjectIdentifier>
                <premis:linkingObjectIdentifierType>preservica_uuid</premis:linkingObjectIdentifierType>
                <premis:linkingObjectIdentifierValue>{preservica_uuid}</premis:linkingObjectIdentifierValue>
            </premis:linkingObjectIdentifier>
        </premis:event>
    </premis:premis>'''.format(preservica_uuid=preservica_uuid, rights_uuid=rights_uuid, rights_basis=rights_basis, rights_status=rights_status, rights_jurisdiction=rights_jurisdiction, rights_date=rights_date, rights_note=rights_note, rights_doc_text=rights_doc_text, rights_doc_uri=rights_doc_uri, event_1_uuid=event_1_uuid, event_1_type=event_1_type, event_1_datetime=event_1_datetime, event_1_details=event_1_details, event_1_agent=event_1_agent)
        premis_path = os.path.join(proj_path, preservica_uuid + '.xml')
        with open(premis_path, 'w') as premis_hand:
            premis_hand.write(premis)
        asset = client.asset(preservica_uuid)
        with open(premis_path, 'r', encoding="utf-8") as md:
            asset = client.add_metadata(asset, "http://www.loc.gov/premis/v3", md)
            print('Appended PREMIS metadata to {}'.format(preservica_uuid))
        os.remove(premis_path)
    fhand.close()
    print('appended {} PREMIS files'.format(count))
# premis_generator()

#this function generates a dictionary of filename:hash values for both the Droid file manifest
#and the ingested Preservica assets (using the Preservica API) and then compares the two to
#ensure they are identical, and provides a report if that is not the case
#TODO update Droid CSV file name (line 515)
def quality_control():
    print('---STARTING QA---')
    asset_count = 0
    print('----MAKING DROID DICTIONARY---')
    droiddict = dict()
    with open('***UPDATE TO DROID CSV FILE***', newline = '') as csvfile:
        reader = csv.reader(csvfile, delimiter = ',', quotechar = '"')
        for row in reader: 
            if 'File' in row[8]:
                droiddict[row[4]] = row[12]
    print('---MAKING PRESERVICA DICTIONARY----')
    client = EntityAPI()
    preservicalist = list()
    iterrow = 2
    for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maxrow, max_col = maxcol):
        for cell in row:
            preservicalist.append(ws.cell(row = iterrow, column = prescol).value)
        iterrow += 1
    preservicadict = dict()
    for reference in preservicalist:
        asset = client.asset(reference)
        asset_count += 1
        for representation in client.representations(asset):
            for content_object in client.content_objects(representation):
                for generation in client.generations(content_object):
                    for bitstream in generation.bitstreams:
                        filename = bitstream.filename 
                        for algorithm,value in bitstream.fixity.items():
                            preservicadict[filename] = value    
    print('----COMPARING DICTIONARIES----')            
    diff = DeepDiff(preservicadict, droiddict, verbose_level=2)
    if len(diff) == 0:
        print('QUALITY ASSURANCE PASSED')
    else:
        print(diff)
    # only used for troubleshoooting if comparing dictionaries fails
    # print('----DROID DICTIONARY----')
    # print(droiddict)
    # print('----PRESERVICA DICTIONARY----')
    # print(preservicadict)
# quality_control()

#----------------------------------------------------------------------------------------------
#Functions grouped below for convenience with indicators of interstitial tasks that need run
#----------------------------------------------------------------------------------------------

#TODO Update project path
#TODO update spreadsheet information

def pre_work():
    work_order_cleanup()
    create_container()
# pre_work()

#TODO convert TIFFs to JPEGs

def pax_prep():
    folder_ds_files()
    img_2_pdf()
    representation_preservation_access()
# pax_prep()

#TODO run Droid report

def create_pax_opex():
    create_pax()
    pax_metadata()
    ao_opex_metadata()
    opex_container_metadata()
# create_pax_opex()

#TODO ingest resources into Preservica

# move_opex_aspace()

#TODO run Link Preservica to ASpace

# move_aspace_trash()

#TODO prepare PREMIS metadata CSV sheet

def qc_id_premis():
    dps_identifier()
    premis_generator()
    quality_control()
# qc_id_premis()
